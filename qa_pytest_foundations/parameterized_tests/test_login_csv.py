# ----------------------------------------------------------------------------------------------------------------------------
#To simulate login tests with multiple username-password combinations, where test data is stored in an Excel file.
#This simulates real-world data-driven testing (DDT) in QA automation.
# ----------------------------------------------------------------------------------------------------------------------------

#Importing required libraries
import pytest                      # Pytest for test framework and parametrize
from openpyxl import load_workbook # openpyxl to read Excel (.xlsx) files
import os                          # os for handling file paths

#Helper function to load test data from Excel
def load_login_data_from_excel():
    test_data = []  # List to store tuples of (username, password, expected)

    # Get the full path to the Excel file, assuming it's in the same folder as this script
    excel_path = os.path.join(os.path.dirname(__file__), 'login_test_data.xlsx')

    # Load the workbook and select the active sheet
    workbook = load_workbook(excel_path)
    sheet = workbook.active  # Use the first sheet in the workbook

    # Iterate over rows, starting from row 2 to skip headers
    for row in sheet.iter_rows(min_row=2, values_only=True):
        username, password, expected = row
        test_data.append((username, password, expected))  # Add tuple to list

    return test_data  # Return the full list of test cases

#Dummy login function simulating real system authentication
def login(username, password):
    # Dictionary of valid users for simulation
    valid_users = {
        "admin": "admin123",
        "testuser": "test123"
    }

# Simulate login logic
#Is the username(from csv) present in the dictionary valid_users? and If yes, does the password of username in dictionary match the expected value in csv?
    if username in valid_users and valid_users[username] == password:
        return "success"
    return "failure"

#Parameterized test using data from Excel
@pytest.mark.parametrize("username, password, expected", load_login_data_from_excel())
def test_login(username, password, expected):
    # Act: call the login function with provided input
    result = login(username, password)

    # Assert: check if result matches expected outcome from Excel
    assert result == expected
